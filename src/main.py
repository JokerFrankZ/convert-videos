from __future__ import annotations

import io
import re
import sys
import time
from contextlib import redirect_stdout
from pathlib import Path
from typing import Iterable, List, NamedTuple, Optional

from PySide6.QtCore import Qt, QMimeData, QThread, QRegularExpression, Signal
from PySide6.QtGui import (
    QColor,
    QDragEnterEvent,
    QDropEvent,
    QFont,
    QSyntaxHighlighter,
    QTextCharFormat,
    QTextCursor,
)
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from converter import (
    DEFAULT_EXPORT_FORMATS,
    ConversionProgress,
    ConversionRequest,
    ConversionTask,
    ConverterError,
    ControlSignals,
    convert_files,
    probe_animated_image_metadata,
    probe_image_metadata,
    probe_video_metadata,
)
from openpyxl import load_workbook

from process_excel import (
    DEFAULT_TEMPLATE,
    RESULT_HEADER,
    process_excel as execute_excel_template,
)


QUALITY_OPTIONS = [
    ("低", "low"),
    ("中等", "medium"),
    ("高", "high"),
    ("超高", "ultra"),
]

SCALE_MODE_OPTIONS = [
    ("居中裁切适配", "center_crop"),
    ("拉伸适配", "stretch"),
    ("强制保持原始宽高比", "force_aspect"),
]

VIDEO_EXTENSIONS = {".mp4", ".mov", ".m4v", ".mpg", ".mpeg"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}
ANIMATED_EXTENSIONS = {".gif"}  # GIF 和 APNG(.png) 动画格式
SUPPORTED_EXTENSIONS = VIDEO_EXTENSIONS | IMAGE_EXTENSIONS | ANIMATED_EXTENSIONS


class SequenceInfo(NamedTuple):
    pattern: str
    first_frame: Path
    extension: str
    start_number: int
    frame_count: int
    prefix: str
    padding: int


class ConversionWorker(QThread):
    progress_signal = Signal(object)
    log_signal = Signal(str)
    error_signal = Signal(str)
    finished_signal = Signal()
    paused_signal = Signal(bool)

    def __init__(self, request: ConversionRequest):
        super().__init__()
        self._request = request
        self._signals = ControlSignals()
        self._signals.pause_event.set()
        self._pause_btn_state = False

    def pause(self) -> None:
        self._signals.request_pause()
        self.paused_signal.emit(True)

    def resume(self) -> None:
        self._signals.request_resume()
        self.paused_signal.emit(False)

    def cancel(self, reason: str | None = None) -> None:
        self._signals.request_cancel(reason)

    def run(self) -> None:  # pragma: no cover - UI thread
        try:
            self._request.signals = self._signals
            convert_files(
                self._request,
                progress=self._forward_progress,
                log=self.log_signal.emit,
            )
        except ConversionCancelled:
            self.log_signal.emit("转换已被终止")
        except ConverterError as exc:
            self.error_signal.emit(str(exc))
        except Exception as exc:  # noqa: BLE001
            self.error_signal.emit(str(exc))
        finally:
            self.finished_signal.emit()

    def _forward_progress(self, progress: ConversionProgress) -> None:
        while (
            not self._signals.pause_event.is_set()
            and not self._signals.cancel_event.is_set()
        ):
            time.sleep(0.1)
        if self._signals.cancel_event.is_set():
            raise ConversionCancelled
        self.progress_signal.emit(progress)


class ConversionCancelled(Exception):
    """用户取消转换时抛出的异常。"""


class _ExcelLogStream(io.TextIOBase):
    def __init__(self, emit_line):
        super().__init__()
        self._emit_line = emit_line
        self._buffer = ""

    def write(self, text):  # type: ignore[override]
        if not text:
            return 0
        self._buffer += text.replace("\r\n", "\n").replace("\r", "\n")
        while "\n" in self._buffer:
            line, self._buffer = self._buffer.split("\n", 1)
            if line:
                self._emit_line(line)
        return len(text)

    def flush(self):  # type: ignore[override]
        if self._buffer:
            self._emit_line(self._buffer)
            self._buffer = ""


class ExcelWorker(QThread):
    log_signal = Signal(str)
    success_signal = Signal(str)
    error_signal = Signal(str)
    finished_signal = Signal()

    def __init__(
        self,
        input_path: str,
        output_path: Optional[str],
        template_text: Optional[str],
        template_path: Optional[str] = None,
    ):
        super().__init__()
        self._input_path = input_path
        self._output_path = output_path
        self._template_text = template_text
        self._template_path = template_path

    def run(self) -> None:  # pragma: no cover - UI thread
        stream = _ExcelLogStream(self.log_signal.emit)
        try:
            with redirect_stdout(stream):
                execute_excel_template(
                    self._input_path,
                    self._output_path or None,
                    template_file=self._template_path or None,
                    template_text=self._template_text or None,
                )
            stream.flush()
            self.success_signal.emit("✅ Excel 模板填充完成")
        except SystemExit as exc:
            stream.flush()
            code = exc.code
            if code not in (None, 0):
                message = str(code) if isinstance(code, str) else "处理失败，请检查输入"
                self.error_signal.emit(message)
            else:
                self.success_signal.emit("✅ Excel 模板填充完成")
        except Exception as exc:  # noqa: BLE001
            stream.flush()
            self.error_signal.emit(str(exc))
        finally:
            self.finished_signal.emit()


class HTMLTemplateHighlighter(QSyntaxHighlighter):
    """HTML模板语法高亮器，支持HTML标签和变量占位符高亮。"""

    def __init__(self, document):
        super().__init__(document)

        # 定义高亮格式
        # 变量占位符格式 {variable}
        self.variable_format = QTextCharFormat()
        self.variable_format.setForeground(QColor("#FF6B6B"))  # 红色
        self.variable_format.setFontWeight(QFont.Bold)

        # HTML标签格式 <tag>
        self.tag_format = QTextCharFormat()
        self.tag_format.setForeground(QColor("#4ECDC4"))  # 青色
        self.tag_format.setFontWeight(QFont.Bold)

        # HTML属性名格式
        self.attribute_name_format = QTextCharFormat()
        self.attribute_name_format.setForeground(QColor("#95E1D3"))  # 浅青色

        # HTML属性值格式
        self.attribute_value_format = QTextCharFormat()
        self.attribute_value_format.setForeground(QColor("#F38181"))  # 浅红色

        # 定义正则表达式规则
        self.highlighting_rules = []

        # 变量占位符规则: {variable_name}
        self.highlighting_rules.append((
            QRegularExpression(r"\{[^{}]+\}"),
            self.variable_format
        ))

        # HTML标签规则: <tag> </tag> <tag/>
        self.highlighting_rules.append((
            QRegularExpression(r"</?[a-zA-Z][a-zA-Z0-9]*(?:\s|/?>)"),
            self.tag_format
        ))

        # HTML标签结束符 >
        self.highlighting_rules.append((
            QRegularExpression(r"/?>"),
            self.tag_format
        ))

        # HTML属性名规则
        self.highlighting_rules.append((
            QRegularExpression(r'\b[a-zA-Z-]+(?==)'),
            self.attribute_name_format
        ))

        # HTML属性值规则（双引号）
        self.highlighting_rules.append((
            QRegularExpression(r'"[^"]*"'),
            self.attribute_value_format
        ))

        # HTML属性值规则（单引号）
        self.highlighting_rules.append((
            QRegularExpression(r"'[^']*'"),
            self.attribute_value_format
        ))

    def highlightBlock(self, text: str) -> None:
        """对文本块应用语法高亮。"""
        for pattern, format_style in self.highlighting_rules:
            match_iterator = pattern.globalMatch(text)
            while match_iterator.hasNext():
                match = match_iterator.next()
                self.setFormat(
                    match.capturedStart(),
                    match.capturedLength(),
                    format_style
                )


class FileDropLineEdit(QLineEdit):
    def __init__(
        self,
        filters: Optional[Iterable[str]] = None,
        parent: QWidget | None = None,
    ):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self._filters = tuple(f.lower() for f in filters) if filters else ()

    def dragEnterEvent(self, event: QDragEnterEvent) -> None:  # type: ignore[override]
        if self._has_valid_urls(event.mimeData()):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event: QDragEnterEvent) -> None:  # type: ignore[override]
        if self._has_valid_urls(event.mimeData()):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event: QDropEvent) -> None:  # type: ignore[override]
        if not event.mimeData().hasUrls():
            event.ignore()
            return
        for url in event.mimeData().urls():
            path = Path(url.toLocalFile())
            if not path:
                continue
            if self._filters and path.suffix.lower() not in self._filters:
                continue
            self.setText(str(path))
            event.acceptProposedAction()
            return
        event.ignore()

    def _has_valid_urls(self, mime_data: QMimeData) -> bool:
        if not mime_data.hasUrls():
            return False
        for url in mime_data.urls():
            path = Path(url.toLocalFile())
            if not path:
                continue
            if not self._filters or path.suffix.lower() in self._filters:
                return True
        return False


class MainWindow(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("视频转换器")
        self.resize(800, 600)
        self.setAcceptDrops(True)

        self._tab_widget = QTabWidget()

        self._files_list = QListWidget()
        self._output_edit = QLineEdit()
        self._width_edit = QLineEdit()
        self._height_edit = QLineEdit()
        self._fps_edit = QLineEdit()
        self._quality_combo = QComboBox()
        for label, value in QUALITY_OPTIONS:
            self._quality_combo.addItem(label, userData=value)
        self._quality_combo.setCurrentIndex(1)  # 默认"中等"

        self._scale_mode_combo = QComboBox()
        for label, value in SCALE_MODE_OPTIONS:
            self._scale_mode_combo.addItem(label, userData=value)
        self._scale_mode_combo.setCurrentIndex(0)  # 默认"居中裁切适配"

        self._export_gif = QCheckBox("GIF")
        self._export_apng = QCheckBox("APNG")
        self._export_png_sequence = QCheckBox("PNG 序列")

        self._export_gif.setChecked(True)
        self._export_apng.setChecked(True)

        self._log_view = QTextEdit()
        self._log_view.setReadOnly(True)
        self._overall_progress = QProgressBar()
        self._current_progress = QProgressBar()
        self._elapsed_label = QLabel("总耗时: --")
        for bar in (self._overall_progress, self._current_progress):
            bar.setRange(0, 100)
            bar.setValue(0)

        self._add_files_btn = QPushButton("添加视频")
        self._add_folder_btn = QPushButton("添加文件夹")
        self._clear_files_btn = QPushButton("清空")
        self._browse_output_btn = QPushButton("选择输出目录")
        self._start_btn = QPushButton("开始转换")
        self._pause_btn = QPushButton("暂停")
        self._resume_btn = QPushButton("继续")
        self._cancel_btn = QPushButton("终止")
        self._start_btn.setEnabled(False)
        self._pause_btn.setEnabled(True)
        self._resume_btn.setEnabled(False)
        self._cancel_btn.setEnabled(True)

        self._worker: ConversionWorker | None = None
        self._file_set: set[str] = set()
        self._tasks: list[ConversionTask] = []
        self._defaults_applied = False
        self._start_time: float | None = None
        self._was_cancelled: bool = False

        excel_filters = (".xlsx", ".xlsm", ".xltx", ".xltm")
        self._excel_input_edit = FileDropLineEdit(excel_filters)
        self._excel_output_edit = FileDropLineEdit(excel_filters)
        self._excel_input_btn = QPushButton("选择输入文件")
        self._excel_output_btn = QPushButton("选择输出文件")
        self._excel_run_btn = QPushButton("开始处理")
        self._excel_log_view = QTextEdit()
        self._excel_log_view.setReadOnly(True)
        self._excel_run_btn.setEnabled(False)
        self._excel_output_edit.setPlaceholderText("留空则覆盖输入文件")
        self._excel_template_edit = QTextEdit()
        self._excel_template_edit.setPlaceholderText("留空使用默认模板")
        self._excel_template_edit.setMinimumHeight(160)
        # 设置等宽字体以提升代码可读性
        code_font = QFont("Menlo")
        code_font.setStyleHint(QFont.Monospace)
        code_font.setPointSize(12)
        self._excel_template_edit.setFont(code_font)
        # 应用语法高亮器
        self._excel_template_highlighter = HTMLTemplateHighlighter(
            self._excel_template_edit.document()
        )
        self._excel_template_import_btn = QPushButton("导入模板")
        self._excel_template_reset_btn = QPushButton("恢复默认模板")
        self._excel_variables_list = QListWidget()
        self._excel_variables_hint = QLabel("双击变量名可插入模板占位符。")
        self._excel_variables_hint.setWordWrap(True)
        self._excel_variables_list.addItem("请选择 Excel 输入文件以加载变量")

        self._excel_worker: ExcelWorker | None = None
        self._excel_template_path: Optional[str] = None
        self._excel_template_updating = False

        self._setup_tabs()
        self._bind_events()

    def _setup_tabs(self) -> None:
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(self._tab_widget)
        self._tab_widget.addTab(self._create_video_tab(), "视频转换")
        self._tab_widget.addTab(self._create_excel_tab(), "模板填充")

    def _create_video_tab(self) -> QWidget:
        tab = QWidget()
        main_layout = QVBoxLayout(tab)

        files_group = QGroupBox("视频文件")
        files_layout = QVBoxLayout()
        files_button_layout = QHBoxLayout()
        files_button_layout.addWidget(self._add_files_btn)
        files_button_layout.addWidget(self._add_folder_btn)
        files_button_layout.addWidget(self._clear_files_btn)

        files_layout.addLayout(files_button_layout)
        files_layout.addWidget(self._files_list)
        files_group.setLayout(files_layout)

        settings_group = QGroupBox("参数设置")
        settings_layout = QGridLayout()
        settings_layout.addWidget(QLabel("输出目录"), 0, 0)
        settings_layout.addWidget(self._output_edit, 0, 1)
        settings_layout.addWidget(self._browse_output_btn, 0, 2)
        settings_layout.addWidget(QLabel("宽度"), 1, 0)
        settings_layout.addWidget(self._width_edit, 1, 1)
        settings_layout.addWidget(QLabel("高度"), 1, 2)
        settings_layout.addWidget(self._height_edit, 1, 3)
        settings_layout.addWidget(QLabel("帧率"), 2, 0)
        settings_layout.addWidget(self._fps_edit, 2, 1)
        settings_layout.addWidget(QLabel("质量"), 2, 2)
        settings_layout.addWidget(self._quality_combo, 2, 3)
        settings_layout.addWidget(QLabel("裁切模式"), 3, 0)
        settings_layout.addWidget(self._scale_mode_combo, 3, 1, 1, 3)

        export_layout = QHBoxLayout()
        export_layout.addWidget(QLabel("导出格式"))
        export_layout.addWidget(self._export_gif)
        export_layout.addWidget(self._export_apng)
        export_layout.addWidget(self._export_png_sequence)
        export_layout.addStretch(1)

        settings_layout.addLayout(export_layout, 4, 0, 1, 4)
        settings_group.setLayout(settings_layout)

        log_group = QGroupBox("日志")
        log_layout = QVBoxLayout()
        log_layout.addWidget(QLabel("整体进度"))
        log_layout.addWidget(self._overall_progress)
        log_layout.addWidget(QLabel("当前文件进度"))
        log_layout.addWidget(self._current_progress)
        log_layout.addWidget(self._elapsed_label)
        log_layout.addWidget(self._log_view)
        log_group.setLayout(log_layout)

        control_layout = QHBoxLayout()
        control_layout.addWidget(self._start_btn)
        control_layout.addWidget(self._pause_btn)
        control_layout.addWidget(self._resume_btn)
        control_layout.addWidget(self._cancel_btn)
        control_layout.addStretch()

        main_layout.addWidget(files_group)
        main_layout.addWidget(settings_group)
        main_layout.addWidget(log_group)
        main_layout.addLayout(control_layout)

        return tab

    def _create_excel_tab(self) -> QWidget:
        tab = QWidget()
        main_layout = QVBoxLayout(tab)

        form_group = QGroupBox("Excel 文件")
        form_layout = QGridLayout()
        form_layout.addWidget(QLabel("输入文件"), 0, 0)
        form_layout.addWidget(self._excel_input_edit, 0, 1)
        form_layout.addWidget(self._excel_input_btn, 0, 2)
        form_layout.addWidget(QLabel("输出文件"), 1, 0)
        form_layout.addWidget(self._excel_output_edit, 1, 1)
        form_layout.addWidget(self._excel_output_btn, 1, 2)
        form_layout.addWidget(QLabel("模板内容"), 2, 0, 1, 1)
        form_layout.addWidget(self._excel_template_edit, 2, 1, 3, 1)
        buttons_layout = QVBoxLayout()
        buttons_layout.addWidget(self._excel_template_import_btn)
        buttons_layout.addWidget(self._excel_template_reset_btn)
        buttons_layout.addStretch(1)
        form_layout.addLayout(buttons_layout, 2, 2, 3, 1)
        form_group.setLayout(form_layout)

        variables_group = QGroupBox("可用模板变量")
        variables_layout = QVBoxLayout()
        variables_layout.addWidget(self._excel_variables_list)
        variables_layout.addWidget(self._excel_variables_hint)
        variables_group.setLayout(variables_layout)

        log_group = QGroupBox("日志")
        log_layout = QVBoxLayout()
        log_layout.addWidget(self._excel_log_view)
        log_group.setLayout(log_layout)

        control_layout = QHBoxLayout()
        control_layout.addStretch()
        control_layout.addWidget(self._excel_run_btn)

        main_layout.addWidget(form_group)
        main_layout.addWidget(variables_group)
        main_layout.addWidget(log_group)
        main_layout.addLayout(control_layout)
        main_layout.addStretch(1)

        return tab

    def _bind_events(self) -> None:
        self._add_files_btn.clicked.connect(self._on_add_files)  # type: ignore[arg-type]
        self._add_folder_btn.clicked.connect(self._on_add_folder)  # type: ignore[arg-type]
        self._clear_files_btn.clicked.connect(self._on_clear_files)  # type: ignore[arg-type]
        self._browse_output_btn.clicked.connect(self._on_browse_output)  # type: ignore[arg-type]
        self._start_btn.clicked.connect(self._on_start)  # type: ignore[arg-type]
        self._pause_btn.clicked.connect(self._on_pause)  # type: ignore[arg-type]
        self._resume_btn.clicked.connect(self._on_resume)  # type: ignore[arg-type]
        self._cancel_btn.clicked.connect(self._on_cancel)  # type: ignore[arg-type]
        self._excel_input_btn.clicked.connect(self._on_excel_browse_input)  # type: ignore[arg-type]
        self._excel_output_btn.clicked.connect(self._on_excel_browse_output)  # type: ignore[arg-type]
        self._excel_template_import_btn.clicked.connect(self._on_excel_import_template)  # type: ignore[arg-type]
        self._excel_template_reset_btn.clicked.connect(self._on_excel_reset_template)  # type: ignore[arg-type]
        self._excel_run_btn.clicked.connect(self._on_excel_run)  # type: ignore[arg-type]
        self._excel_variables_list.itemDoubleClicked.connect(self._on_excel_variable_double_clicked)  # type: ignore[arg-type]
        self._excel_input_edit.textChanged.connect(self._on_excel_input_changed)  # type: ignore[arg-type]
        self._excel_output_edit.textChanged.connect(self._update_excel_run_state)  # type: ignore[arg-type]
        self._excel_template_edit.textChanged.connect(self._on_excel_template_changed)  # type: ignore[arg-type]

    def _on_add_files(self) -> None:  # pragma: no cover - UI
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "选择媒体文件",
            "",
            "媒体文件 (*.mp4 *.mov *.m4v *.mpg *.mpeg *.gif *.png)",
        )

        self._add_paths(Path(path) for path in paths)

    def _on_clear_files(self) -> None:
        self._files_list.clear()
        self._file_set.clear()
        self._tasks.clear()
        self._defaults_applied = False
        self._width_edit.clear()
        self._height_edit.clear()
        self._fps_edit.clear()
        self._overall_progress.setValue(0)
        self._current_progress.setValue(0)
        self._elapsed_label.setText("总耗时: --")
        self._set_controls_stopped()
        self._update_start_state()

    def _on_browse_output(self) -> None:
        directory = QFileDialog.getExistingDirectory(self, "选择输出目录")
        if directory:
            self._output_edit.setText(directory)
            self._update_start_state()

    def _on_add_folder(self) -> None:
        directory = QFileDialog.getExistingDirectory(self, "选择视频所在文件夹")
        if not directory:
            return

        folder_path = Path(directory)
        paths = (path for path in folder_path.rglob("*") if path.is_file())
        self._add_paths(paths)

    def _update_start_state(self) -> None:
        has_files = self._files_list.count() > 0
        has_output = bool(self._output_edit.text())
        self._start_btn.setEnabled(has_files and has_output and self._worker is None)

    def _append_log(self, message: str) -> None:
        self._log_view.append(message)
        self._log_view.moveCursor(QTextCursor.End)

    def _on_progress(self, progress: ConversionProgress) -> None:
        overall_percent = int(min(max(progress.overall_progress * 100, 0), 100))
        task_percent = int(min(max(progress.task_progress * 100, 0), 100))
        self._overall_progress.setValue(overall_percent)
        self._current_progress.setValue(task_percent)
        self._append_log(
            f"[{progress.task_index}/{progress.total_tasks}] {progress.task_name} - {progress.stage} ({task_percent}%)"
        )

    def _on_start(self) -> None:
        try:
            request = self._build_request()
        except ConverterError as exc:
            QMessageBox.warning(self, "参数错误", str(exc))
            return

        self._log_view.clear()
        self._append_log("开始转换...")
        self._start_btn.setEnabled(False)
        self._pause_btn.setEnabled(True)
        self._resume_btn.setEnabled(False)
        self._cancel_btn.setEnabled(True)

        self._start_time = time.perf_counter()
        self._was_cancelled = False

        self._overall_progress.setValue(0)
        self._current_progress.setValue(0)
        self._worker = ConversionWorker(request)
        self._worker.progress_signal.connect(self._on_progress)
        self._worker.log_signal.connect(self._append_log)
        self._worker.error_signal.connect(self._on_error)
        self._worker.finished_signal.connect(self._on_finished)
        self._worker.paused_signal.connect(self._on_worker_paused)
        self._worker.start()
        self._pause_btn.setEnabled(True)
        self._resume_btn.setEnabled(False)
        self._cancel_btn.setEnabled(True)

    def _on_error(self, message: str) -> None:
        self._append_log(f"❌ {message}")
        if not self._was_cancelled:
            QMessageBox.critical(self, "转换失败", message)
        self._set_controls_stopped()

    def _on_finished(self) -> None:
        elapsed = "--"
        if self._start_time is not None:
            elapsed_seconds = time.perf_counter() - self._start_time
            elapsed = f"{elapsed_seconds:.2f} 秒"
        self._elapsed_label.setText(f"总耗时: {elapsed}")
        if self._was_cancelled:
            self._append_log("⚠️ 转换已终止")
        else:
            self._append_log("✅ 转换完成")
        self._worker = None
        self._set_controls_stopped()
        self._was_cancelled = False

    def _build_request(self) -> ConversionRequest:
        output_dir = Path(self._output_edit.text()).expanduser()
        if not output_dir:
            raise ConverterError("请指定输出目录")

        try:
            width = int(self._width_edit.text())
            height = int(self._height_edit.text())
            fps = float(self._fps_edit.text())
        except ValueError as exc:
            raise ConverterError("宽度、高度、帧率必须是数字") from exc

        if width <= 0 or height <= 0 or fps <= 0:
            raise ConverterError("宽度、高度、帧率必须大于 0")

        quality = self._quality_combo.currentData()
        scale_mode = self._scale_mode_combo.currentData()

        return ConversionRequest(
            tasks=self._tasks,
            output_dir=output_dir,
            width=width,
            height=height,
            fps=fps,
            quality=quality,
            scale_mode=scale_mode,
            export_formats=self._gather_export_formats(),
        )

    def _gather_export_formats(self) -> tuple[str, ...]:
        formats: list[str] = []
        if self._export_gif.isChecked():
            formats.append("gif")
        if self._export_apng.isChecked():
            formats.append("apng")
        if self._export_png_sequence.isChecked():
            formats.append("png_sequence")
        if not formats:
            # 如果没有勾选任何格式，根据源文件格式推断
            source_formats = set()
            for task in self._tasks:
                if task.source_format:
                    source_formats.add(task.source_format)

            # 根据源格式映射到导出格式
            for src_fmt in source_formats:
                if src_fmt == "gif":
                    formats.append("gif")
                elif src_fmt == "apng":
                    formats.append("apng")
                elif src_fmt in ("video", "image_sequence"):
                    # 视频和图片序列默认导出为 GIF 和 APNG
                    if "gif" not in formats:
                        formats.append("gif")
                    if "apng" not in formats:
                        formats.append("apng")

            # 如果还是没有格式（比如没有任务），使用默认格式
            if not formats:
                formats.extend(DEFAULT_EXPORT_FORMATS)

        return tuple(dict.fromkeys(formats))

    def _add_paths(self, paths: Iterable[Path]) -> None:
        new_tasks: list[ConversionTask] = []
        for path in paths:
            normalized = path.resolve()
            if not normalized.exists():
                continue

            suffix = normalized.suffix.lower()
            key = str(normalized)

            if suffix in VIDEO_EXTENSIONS:
                if key in self._file_set:
                    continue
                try:
                    width, height, fps, frames, duration = probe_video_metadata(
                        normalized
                    )
                except ConverterError as exc:
                    self._append_log(f"⚠️ 无法读取 {normalized.name} 的参数：{exc}")
                    continue
                task = ConversionTask(
                    display_name=normalized.name,
                    source=normalized,
                    output_stem=normalized.stem,
                    source_format="video",
                    total_frames=frames,
                    duration_ms=duration,
                )
                self._file_set.add(key)
                new_tasks.append(task)
                continue

            # 处理 GIF 动画
            if suffix in ANIMATED_EXTENSIONS:
                if key in self._file_set:
                    continue
                try:
                    width, height, fps, frames, duration = probe_animated_image_metadata(
                        normalized
                    )
                except ConverterError as exc:
                    self._append_log(f"⚠️ 无法读取 {normalized.name} 的参数：{exc}")
                    continue
                task = ConversionTask(
                    display_name=normalized.name,
                    source=normalized,
                    output_stem=normalized.stem,
                    source_format="gif",
                    total_frames=frames,
                    duration_ms=duration,
                )
                self._file_set.add(key)
                new_tasks.append(task)
                continue

            if suffix in IMAGE_EXTENSIONS:
                # 尝试检测是否为 APNG（动画 PNG）
                if suffix == ".png":
                    try:
                        width, height, fps, frames, duration = probe_animated_image_metadata(
                            normalized
                        )
                        # 如果成功获取帧数且大于1，说明是 APNG
                        if frames and frames > 1:
                            if key in self._file_set:
                                continue
                            task = ConversionTask(
                                display_name=normalized.name,
                                source=normalized,
                                output_stem=normalized.stem,
                                source_format="apng",
                                total_frames=frames,
                                duration_ms=duration,
                            )
                            self._file_set.add(key)
                            new_tasks.append(task)
                            continue
                    except ConverterError:
                        # 如果探测失败，按照普通图片处理
                        pass

                sequence = self._detect_sequence(normalized)
                if sequence:
                    sequence_key = sequence.pattern
                    if sequence_key in self._file_set:
                        continue
                    task = self._build_sequence_task(sequence)
                    if task:
                        task.source_format = "image_sequence"
                        task.frame_count = sequence.frame_count
                        task.duration_ms = sequence.frame_count * 150
                        self._file_set.add(sequence_key)
                        new_tasks.append(task)
                    continue

                if key in self._file_set:
                    continue
                task = ConversionTask(
                    display_name=normalized.name,
                    source=normalized,
                    output_stem=normalized.stem,
                    source_format="image_sequence",
                    is_sequence=True,
                    sequence_pattern=str(normalized),
                    frame_extension=suffix,
                    frame_count=1,
                    first_frame=normalized,
                    start_number=0,
                    duration_ms=150,
                )
                self._file_set.add(key)
                new_tasks.append(task)

        if not new_tasks:
            self._append_log("未添加新的可转换文件")
            return

        for task in new_tasks:
            item = QListWidgetItem(task.display_name)
            tooltip_parts = [f"输出: {task.output_stem}"]
            tooltip_parts.append(task.sequence_pattern or str(task.source))
            if task.total_frames:
                tooltip_parts.append(f"帧数: {task.total_frames}")
            if task.duration_ms:
                tooltip_parts.append(f"时长: {task.duration_ms / 1000:.2f}s")
            item.setToolTip(" | ".join(tooltip_parts))
            item.setData(Qt.ItemDataRole.UserRole, task)
            self._files_list.addItem(item)
            self._tasks.append(task)

        self._update_start_state()
        self._apply_defaults_if_needed()
        self._append_log(f"已添加 {len(new_tasks)} 个任务")

    def _on_excel_browse_input(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择 Excel 文件",
            "",
            "Excel 文件 (*.xlsx *.xlsm *.xltx *.xltm);;所有文件 (*)",
        )
        if path:
            self._excel_input_edit.setText(path)

    def _on_excel_browse_output(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self,
            "选择输出 Excel 文件",
            "",
            "Excel 文件 (*.xlsx);;所有文件 (*)",
        )
        if path:
            self._excel_output_edit.setText(path)

    def _on_excel_import_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择 HTML 模板",
            "",
            "HTML 模板 (*.html *.htm *.txt);;所有文件 (*)",
        )
        if path:
            try:
                with open(path, "r", encoding="utf-8") as file:
                    content = file.read()
            except Exception as exc:  # noqa: BLE001
                QMessageBox.warning(self, "读取失败", f"无法读取模板文件：{exc}")
                return
            self._excel_template_updating = True
            self._excel_template_edit.setPlainText(content)
            self._excel_template_updating = False
            self._excel_template_path = path

    def _on_excel_reset_template(self) -> None:
        self._excel_template_updating = True
        self._excel_template_edit.setPlainText(DEFAULT_TEMPLATE)
        self._excel_template_updating = False
        self._excel_template_path = None

    def _on_excel_template_changed(self) -> None:
        if self._excel_template_updating:
            return
        self._excel_template_path = None

    def _on_excel_run(self) -> None:
        if self._excel_worker is not None:
            return

        input_path = self._excel_input_edit.text().strip()
        output_path = self._excel_output_edit.text().strip() or None
        raw_template_text = self._excel_template_edit.toPlainText()
        template_text = raw_template_text if raw_template_text.strip() else None

        if not input_path:
            QMessageBox.warning(self, "参数错误", "请选择需要处理的 Excel 文件")
            return

        self._excel_log_view.clear()
        self._set_excel_controls_running(True)

        self._excel_worker = ExcelWorker(
            input_path,
            output_path,
            template_text,
            self._excel_template_path,
        )
        self._excel_worker.log_signal.connect(self._append_excel_log)
        self._excel_worker.error_signal.connect(self._on_excel_error)
        self._excel_worker.success_signal.connect(self._on_excel_success)
        self._excel_worker.finished_signal.connect(self._on_excel_finished)
        self._excel_worker.start()

    def _append_excel_log(self, message: str) -> None:
        self._excel_log_view.append(message)
        self._excel_log_view.moveCursor(QTextCursor.End)

    def _on_excel_error(self, message: str) -> None:
        self._append_excel_log(f"❌ {message}")
        QMessageBox.critical(self, "处理失败", message)

    def _on_excel_success(self, message: str) -> None:
        self._append_excel_log(message)
        QMessageBox.information(self, "处理完成", "Excel 模板填充完成")

    def _on_excel_finished(self) -> None:
        self._excel_worker = None
        self._set_excel_controls_running(False)

    def _set_excel_controls_running(self, running: bool) -> None:
        self._excel_run_btn.setEnabled(not running and bool(self._excel_input_edit.text().strip()))
        for widget in (
            self._excel_input_btn,
            self._excel_output_btn,
            self._excel_input_edit,
            self._excel_output_edit,
            self._excel_template_edit,
            self._excel_template_import_btn,
            self._excel_template_reset_btn,
            self._excel_variables_list,
        ):
            widget.setEnabled(not running)

    def _on_excel_input_changed(self) -> None:
        self._update_excel_run_state()
        self._update_excel_variables()

    def _update_excel_variables(self) -> None:
        path = self._excel_input_edit.text().strip()
        self._excel_variables_list.clear()
        if not path:
            self._excel_variables_list.addItem("请先选择 Excel 输入文件")
            return

        file_path = Path(path)
        if not file_path.exists():
            self._excel_variables_list.addItem("⚠️ 文件不存在")
            return

        try:
            workbook = load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            self._excel_variables_list.addItem(f"⚠️ 读取失败: {exc}")
            return

        try:
            sheet = workbook.active
            headers: list[str] = []
            duplicates: set[str] = set()
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row=1, column=col).value
                if value is None:
                    continue
                header = str(value).strip()
                if not header or header == RESULT_HEADER:
                    continue
                if header in headers:
                    duplicates.add(header)
                    continue
                headers.append(header)
        finally:
            workbook.close()

        if not headers:
            self._excel_variables_list.addItem("⚠️ 首行未检测到模板变量")
            return

        for header in headers:
            item = QListWidgetItem(header)
            item.setData(Qt.ItemDataRole.UserRole, header)
            self._excel_variables_list.addItem(item)

        if duplicates:
            note = f"⚠️ 重复变量已忽略: {', '.join(sorted(duplicates))}"
            warning_item = QListWidgetItem(note)
            warning_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            warning_item.setData(Qt.ItemDataRole.UserRole, None)
            self._excel_variables_list.addItem(warning_item)

    def _on_excel_variable_double_clicked(self, item: QListWidgetItem) -> None:
        header = item.data(Qt.ItemDataRole.UserRole)
        if not header:
            return
        cursor = self._excel_template_edit.textCursor()
        cursor.insertText(f"{{{header}}}")
        self._excel_template_edit.setTextCursor(cursor)
        self._excel_template_edit.setFocus()

    def _update_excel_run_state(self) -> None:
        if self._excel_worker is not None:
            return
        has_input = bool(self._excel_input_edit.text().strip())
        self._excel_run_btn.setEnabled(has_input)

    def _detect_sequence(self, frame_path: Path) -> SequenceInfo | None:
        match = re.match(r"(.*?)(\d+)(\.[^.]+)$", frame_path.name)
        if not match:
            return None

        prefix, number_str, extension = match.groups()
        padding = len(number_str)
        directory = frame_path.parent
        regex = re.compile(rf"^{re.escape(prefix)}(\d+){re.escape(extension)}$")

        frames: List[tuple[int, Path]] = []
        for candidate in directory.iterdir():
            if not candidate.is_file():
                continue
            candidate_match = regex.match(candidate.name)
            if not candidate_match:
                continue
            try:
                frame_number = int(candidate_match.group(1))
            except ValueError:
                continue
            frames.append((frame_number, candidate.resolve()))

        if len(frames) < 2:
            return None

        frames.sort(key=lambda item: item[0])
        start_number, first_frame = frames[0]
        frame_count = len(frames)
        pattern = str(directory / f"{prefix}%0{padding}d{extension}")
        prefix_name = (
            prefix.rstrip("_- ")
            or prefix
            or first_frame.parent.name
            or first_frame.stem
        )

        return SequenceInfo(
            pattern=pattern,
            first_frame=first_frame,
            extension=extension,
            start_number=start_number,
            frame_count=frame_count,
            prefix=prefix_name,
            padding=padding,
        )

    def _build_sequence_task(self, sequence: SequenceInfo) -> ConversionTask | None:
        output_stem = (
            sequence.prefix.strip()
            or sequence.first_frame.parent.name
            or sequence.first_frame.stem
        )
        output_stem = output_stem.replace(" ", "_")
        display_name = (
            f"{output_stem}{sequence.extension} 序列 ({sequence.frame_count} 张)"
        )

        return ConversionTask(
            display_name=display_name,
            source=sequence.first_frame,
            output_stem=output_stem,
            is_sequence=True,
            sequence_pattern=sequence.pattern,
            frame_extension=sequence.extension,
            frame_count=sequence.frame_count,
            first_frame=sequence.first_frame,
            start_number=sequence.start_number,
            duration_ms=sequence.frame_count * 150,
        )

    def _current_fps(self) -> float:
        text = self._fps_edit.text().strip()
        try:
            return float(text) if text else 10.0
        except ValueError:
            return 10.0

    def _apply_defaults_if_needed(self) -> None:
        if self._defaults_applied or not self._tasks:
            return

        first_task = self._tasks[0]
        try:
            if first_task.is_sequence and first_task.first_frame:
                width, height = probe_image_metadata(first_task.first_frame)
                fps = self._current_fps()
            elif first_task.source.suffix.lower() in IMAGE_EXTENSIONS:
                width, height = probe_image_metadata(first_task.source)
                fps = self._current_fps()
            else:
                width, height, fps, frames, duration = probe_video_metadata(
                    first_task.source
                )
                first_task.total_frames = frames
                first_task.duration_ms = duration
        except ConverterError as exc:
            self._append_log(f"⚠️ 无法读取 {first_task.display_name} 的参数：{exc}")
            return

        self._width_edit.setText(str(width))
        self._height_edit.setText(str(height))
        self._fps_edit.setText(f"{fps:.3f}".rstrip("0").rstrip("."))
        self._defaults_applied = True

    def showEvent(self, event):  # type: ignore[override]
        super().showEvent(event)
        if not self._output_edit.text():
            default_output = self._default_output_directory()
            self._output_edit.setText(str(default_output))
            self._update_start_state()

    def _default_output_directory(self) -> Path:
        home = Path.home()
        downloads = home / "Downloads"
        if downloads.exists():
            return downloads
        if sys.platform.startswith("win"):
            from ctypes import windll, wintypes, create_unicode_buffer

            CSIDL_PERSONAL = 0x0005
            SHGFP_TYPE_CURRENT = 0
            buf = create_unicode_buffer(wintypes.MAX_PATH)
            result = windll.shell32.SHGetFolderPathW(
                None, CSIDL_PERSONAL, None, SHGFP_TYPE_CURRENT, buf
            )
            if result == 0:
                documents = Path(buf.value)
                candidate = documents.parent / "Downloads"
                if candidate.exists():
                    return candidate
        return home

    # 拖拽处理
    def dragEnterEvent(self, event: QDragEnterEvent) -> None:  # type: ignore[override]
        if self._contains_valid_urls(event.mimeData()):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event: QDropEvent) -> None:  # type: ignore[override]
        urls = event.mimeData().urls()
        paths = []
        for url in urls:
            local_path = Path(url.toLocalFile())
            if local_path.is_dir():
                paths.extend(p for p in local_path.rglob("*") if p.is_file())
            else:
                paths.append(local_path)

        filtered = [p for p in paths if p.suffix.lower() in SUPPORTED_EXTENSIONS]
        if filtered:
            self._add_paths(filtered)
            event.acceptProposedAction()
        else:
            self._append_log("拖拽内容未包含可支持的媒体文件")
            event.ignore()

    def _contains_valid_urls(self, mime_data: QMimeData) -> bool:
        if not mime_data.hasUrls():
            return False
        for url in mime_data.urls():
            path = Path(url.toLocalFile())
            if path.is_dir():
                return True
            if path.suffix.lower() in SUPPORTED_EXTENSIONS:
                return True
        return False

    def _on_pause(self) -> None:
        if self._worker:
            self._worker.pause()
            self._pause_btn.setEnabled(False)
            self._resume_btn.setEnabled(True)

    def _on_resume(self) -> None:
        if self._worker:
            self._worker.resume()
            self._pause_btn.setEnabled(True)
            self._resume_btn.setEnabled(False)

    def _on_cancel(self) -> None:
        if self._worker:
            self._was_cancelled = True
            self._worker.cancel("用户终止了转换")
            self._pause_btn.setEnabled(False)
            self._resume_btn.setEnabled(False)
            self._cancel_btn.setEnabled(False)

    def _set_controls_stopped(self) -> None:
        self._start_btn.setEnabled(bool(self._tasks) and bool(self._output_edit.text()))
        self._pause_btn.setEnabled(False)
        self._resume_btn.setEnabled(False)
        self._cancel_btn.setEnabled(False)

    def _on_worker_paused(self, paused: bool) -> None:
        self._pause_btn.setEnabled(not paused)
        self._resume_btn.setEnabled(paused)
        status = "已暂停" if paused else "继续转换"
        self._append_log(status)


def main() -> int:
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
