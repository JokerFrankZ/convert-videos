#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

"""
读取 Excel 文件，首行作为模板变量定义，使用 HTML 模板生成结果列
"""

import argparse
import os
import sys
import re

import openpyxl

DEFAULT_TEMPLATE = """<p>{desc}</p><p>{name}</p><p>{price}</p><p>【具体价格请咨询商家】</p><p><br></p><p>不只是卖花，更是传递思念与欢喜，让每一份情感都有处安放。</p><p><br></p><p><img alt="图片上传" src="https://jmy-pic.baidu.com/0/pic/-1385074748_501952671_1241745945.jpg" class="fr-fic fr-dii"></p><p><img alt="图片上传" src="https://jmy-pic.baidu.com/0/pic/1850010481_-13527842_504047256.jpg" class="fr-fic fr-dii"></p><p><img alt="图片上传" src="https://jmy-pic.baidu.com/0/pic/-484983576_678918380_-191833468.jpg" class="fr-fic fr-dii"></p>"""
RESULT_HEADER = "生成结果"


def _load_template_from_file(template_path: str) -> str:
    if not os.path.exists(template_path):
        print(f"错误：模板文件 '{template_path}' 不存在！")
        sys.exit(1)
    try:
        with open(template_path, "r", encoding="utf-8") as file:
            return file.read()
    except Exception as exc:  # noqa: BLE001
        print(f"错误：读取模板文件失败: {exc}")
        sys.exit(1)


def _resolve_template(template_text: str | None, template_path: str | None) -> str:
    if template_text is not None:
        trimmed = template_text.strip()
        if trimmed:
            return template_text
    if template_path:
        return _load_template_from_file(template_path)
    return DEFAULT_TEMPLATE


def _build_header_mapping(sheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    max_col = sheet.max_column
    for col in range(1, max_col + 1):
        header_value = sheet.cell(row=1, column=col).value
        if header_value is None:
            continue
        header_str = str(header_value).strip()
        if not header_str or header_str == RESULT_HEADER:
            continue
        if header_str in mapping:
            print(f"警告：变量 '{header_str}' 重复，仅保留第一列。")
            continue
        mapping[header_str] = col
    if not mapping:
        print("错误：首行未定义任何模板变量。")
        sys.exit(1)
    return mapping


_PLACEHOLDER_PATTERN = re.compile(r"{([^{}]+)}")


def _render_template(template: str, values: dict[str, str]) -> str:
    def replacer(match: re.Match) -> str:
        key = match.group(1)
        return values.get(key, "")

    return _PLACEHOLDER_PATTERN.sub(replacer, template)


def process_excel(input_file, output_file=None, template_file=None, template_text=None):
    """
    处理 Excel 文件，替换模板中的占位符

    Args:
        input_file: 输入文件路径
        output_file: 输出文件路径（可选，默认覆盖原文件）
        template_file: HTML 模板文件路径（可选）
        template_text: HTML 模板文本内容（可选，高优先级）
    """
    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误：文件 '{input_file}' 不存在！")
        sys.exit(1)

    # 如果没有指定输出文件，则覆盖原文件
    if output_file is None:
        output_file = input_file

    print(f"正在读取文件: {input_file}")

    try:
        # 读取 Excel 文件
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active

        template_content = _resolve_template(template_text, template_file)
        header_mapping = _build_header_mapping(sheet)

        result_column = sheet.max_column + 1
        sheet.cell(row=1, column=result_column, value=RESULT_HEADER)

        processed_count = 0

        # 遍历数据行（从第2行开始）
        for row in range(2, sheet.max_row + 1):
            values: dict[str, str] = {}
            non_empty = False
            for key, col in header_mapping.items():
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is None:
                    values[key] = ""
                else:
                    non_empty = True
                    values[key] = str(cell_value)

            if not non_empty:
                continue

            result = _render_template(template_content, values)

            sheet.cell(row=row, column=result_column, value=result)
            print(f"处理第 {row} 行: {values}")
            processed_count += 1

        # 保存文件
        workbook.save(output_file)
        print(f"\n✅ 处理完成！")
        print(f"   - 共处理 {processed_count} 行数据")
        print(f"   - 结果已保存到: {output_file}")

    except Exception as e:
        print(f"错误：处理文件时出现异常: {e}")
        sys.exit(1)


def main():
    """主函数，处理命令行参数"""
    parser = argparse.ArgumentParser(
        description="读取 Excel 文件，首行定义模板变量，使用 HTML 模板生成结果列",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  %(prog)s input.xlsx                    # 处理 input.xlsx，结果覆盖原文件
  %(prog)s input.xlsx -o output.xlsx     # 处理 input.xlsx，结果保存到 output.xlsx
  %(prog)s /path/to/doc.xlsx             # 使用完整路径
        """,
    )

    parser.add_argument("input", help="输入 Excel 文件路径")

    parser.add_argument(
        "-o",
        "--output",
        help="输出 Excel 文件路径（可选，默认覆盖输入文件）",
        default=None,
    )

    parser.add_argument(
        "-t",
        "--template",
        help="HTML 模板文件路径（可选，默认使用内置模板）",
        default=None,
    )
    parser.add_argument(
        "--template-text",
        help="直接传入 HTML 模板内容（优先于 --template）",
        default=None,
    )

    args = parser.parse_args()

    # 处理文件
    process_excel(args.input, args.output, args.template, args.template_text)


if __name__ == "__main__":
    main()
